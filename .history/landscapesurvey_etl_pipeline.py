"""
ETL Script for loading Landscape Analysis Excel data into PostgreSQL
Schema: countryprofiles

Rules:
1. Insert sheet names into countries table
2. Rows A1–A3 => Indicators under fixed category "Population and Economy"
3. Rows A4–A69 => If col B empty => indicator category; else => indicator under latest category
"""

import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
from pathlib import Path
from openpyxl import load_workbook
from typing import Optional

# ---------------------------
# CONFIGURATION
# ---------------------------

EXCEL_PATH = Path("C:/Users/Public/Landscape_analysis.xlsx")

# Database connection settings
DB_NAME = "emp_pip"
DB_USER = "muzat"
DB_PASS = "@muzat143!"
DB_HOST = "db-epr-hir-postgresql.postgres.database.azure.com"
DB_PORT = 5432  # int is fine

SURVEY_YEAR = 2024
SCHEMA = "countryprofiles"

# ---------------------------
# DB Functions
# ---------------------------

def connect_db():
    return psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASS,
        host=DB_HOST,
        port=DB_PORT
    )

def reset_tables(cur, schema="public"):
    """Reset all tables in the correct order to avoid foreign key violations"""
    try:
        cur.execute(f"""
            TRUNCATE TABLE {schema}.country_indicators RESTART IDENTITY CASCADE;
            TRUNCATE TABLE {schema}.indicators RESTART IDENTITY CASCADE;
            TRUNCATE TABLE {schema}.indicator_categories RESTART IDENTITY CASCADE;
            TRUNCATE TABLE {schema}.countries RESTART IDENTITY CASCADE;
        """)
        print("✓ All tables reset successfully")
    except Exception as e:
        print(f"Error resetting tables: {e}")
        raise

def upsert_country(cur, name: str) -> int:
    """Insert or update country and return country_id"""
    try:
        cur.execute(
            f"""
            INSERT INTO {SCHEMA}.countries (name)
            VALUES (%s)
            ON CONFLICT (name)
            DO UPDATE SET name = EXCLUDED.name
            RETURNING country_id;
            """,
            (name.strip(),)
        )
        result = cur.fetchone()
        country_id = result[0]
        print(f"✓ Country: {country_id} {name} successfully processed")
        return country_id
    except Exception as e:
        print(f"Error upserting country {name}: {e}")
        raise

def upsert_category(cur, cat_name: str) -> int:
    """Insert or update category and return category_id"""
    try:
        cur.execute(
            f"""
            INSERT INTO {SCHEMA}.indicator_categories (cat_name)
            VALUES (%s)
            ON CONFLICT (cat_name)
            DO UPDATE SET cat_name = EXCLUDED.cat_name
            RETURNING category_id;
            """,
            (cat_name.strip(),)
        )
        result = cur.fetchone()
        category_id = result[0]
        print(f"✓ Category: {category_id} {cat_name} successfully processed")
        return category_id
    except Exception as e:
        print(f"Error upserting category {cat_name}: {e}")
        raise

def upsert_indicator(cur, name: str, category_id: Optional[int] = None) -> int:
    """Insert or update indicator and return indicator_id"""
    try:
        if category_id is not None:
            # First verify the category exists
            cur.execute(f"SELECT 1 FROM {SCHEMA}.indicator_categories WHERE category_id = %s", (category_id,))
            if not cur.fetchone():
                raise ValueError(f"Category ID {category_id} does not exist")
                
            cur.execute(
                f"""
                INSERT INTO {SCHEMA}.indicators (name, category_id)
                VALUES (%s, %s)
                ON CONFLICT (name)
                DO UPDATE SET category_id = EXCLUDED.category_id
                RETURNING indicator_id;
                """,
                (name.strip(), category_id)
            )
        else:
            cur.execute(
                f"""
                INSERT INTO {SCHEMA}.indicators (name)
                VALUES (%s)
                ON CONFLICT (name)
                DO UPDATE SET name = EXCLUDED.name
                RETURNING indicator_id;
                """,
                (name.strip(),)
            )
        result = cur.fetchone()
        indicator_id = result[0]
        print(f"✓ Indicator: {indicator_id} {name} successfully processed")
        return indicator_id
    except Exception as e:
        print(f"Error upserting indicator {name}: {e}")
        raise

def insert_country_indicators(cur, records):
    """Insert country indicators"""
    if not records:
        print("No records to insert")
        return
        
    try:
        sql = f"""
        INSERT INTO {SCHEMA}.country_indicators (country_id, indicator_id, survey_year, survey_response, countryname, indicatorname)
        VALUES %s
        ON CONFLICT (country_id, indicator_id, survey_year)
        DO UPDATE SET survey_response = EXCLUDED.survey_response, 
                      countryname = EXCLUDED.countryname, 
                      indicatorname = EXCLUDED.indicatorname;
        """
        execute_values(cur, sql, records, page_size=100)
        print(f"✓ Successfully inserted {len(records)} country indicator records")
        
    except Exception as e:
        print(f"Error inserting country indicators: {e}")
        raise

# ---------------------------
# Excel Parsing Logic
# ---------------------------

def extract_rows(sheet):
    """
    Extract rows A1:B69 with (row, colA_text, colB_value).
    """
    rows = []
    for row in range(1, 70):
        cell_a = sheet[f"A{row}"]
        cell_b = sheet[f"B{row}"]

        text_a = str(cell_a.value).strip() if cell_a.value else None
        value_b = str(cell_b.value).strip() if cell_b.value else None

        rows.append((row, text_a, value_b))
    return rows

# ---------------------------
# ETL Logic
# ---------------------------

def process_excel(excel_path: Path, conn):
    """Main ETL process with comprehensive error handling"""
    try:
        wb = load_workbook(excel_path, data_only=True)
        cur = conn.cursor()
        
        # Reset tables and commit
        reset_tables(cur, schema=SCHEMA)
        conn.commit()
        print("✓ Database reset completed\n")

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            print(f"\n--- Processing sheet {sheet_idx + 1}: {sheet_name} ---")
            sheet = wb[sheet_name]

            country_name = sheet_name.strip()
            
            # Insert country and commit immediately
            country_id = upsert_country(cur, country_name)
            conn.commit()
            
            # Verify country was actually inserted and get fresh connection state
            verify_cur = conn.cursor()
            verify_cur.execute(f"SELECT country_id, name FROM {SCHEMA}.countries WHERE country_id = %s", (country_id,))
            country_check = verify_cur.fetchone()
            verify_cur.close()
            if not country_check:
                raise ValueError(f"Failed to verify country insertion for {country_name} with ID {country_id}")
            print(f"🔍 Verified country in DB: {country_check}")

            records = []
            detected_categories = ["Population and Economy"]
            detected_indicators = {"Population and Economy": []}

            # 1. Insert fixed category "Population and Economy" and commit
            pop_econ_cat_id = upsert_category(cur, "Population and Economy")
            conn.commit()
            
            # Verify category was inserted
            verify_cur = conn.cursor()
            verify_cur.execute(f"SELECT category_id, cat_name FROM {SCHEMA}.indicator_categories WHERE category_id = %s", (pop_econ_cat_id,))
            cat_check = verify_cur.fetchone()
            verify_cur.close()
            print(f"🔍 Verified category in DB: {cat_check}")

            rows = extract_rows(sheet)

            # First 3 rows → Population and Economy indicators
            for row_idx, text, value in rows[:3]:
                if text:
                    ind_id = upsert_indicator(cur, text, pop_econ_cat_id)
                    conn.commit()  # Commit each indicator
                    
                    # Verify indicator was inserted
                    verify_cur = conn.cursor()
                    verify_cur.execute(f"SELECT indicator_id, name FROM {SCHEMA}.indicators WHERE indicator_id = %s", (ind_id,))
                    ind_check = verify_cur.fetchone()
                    verify_cur.close()
                    if not ind_check:
                        raise ValueError(f"Failed to verify indicator insertion for {text} with ID {ind_id}")
                    
                    records.append((country_id, ind_id, SURVEY_YEAR, value, country_name, text))
                    detected_indicators["Population and Economy"].append((text, value))

            # From A4 onwards → category if colB empty, else indicator under last category
            current_cat_name = None
            current_cat_id = None
            
            for row_idx, text, value in rows[3:]:
                if not text:
                    continue

                if not value:  # treat as category
                    current_cat_id = upsert_category(cur, text)
                    conn.commit()  # Commit each category
                    current_cat_name = text
                    detected_categories.append(text)
                    detected_indicators[current_cat_name] = []
                else:  # indicator
                    if current_cat_id is None:
                        print(f"Warning: Indicator '{text}' has no category assigned, skipping")
                        continue
                        
                    ind_id = upsert_indicator(cur, text, current_cat_id)
                    conn.commit()  # Commit each indicator
                    records.append((country_id, ind_id, SURVEY_YEAR, value, country_name, text))
                    if current_cat_name:
                        detected_indicators[current_cat_name].append((text, value))

            # Process country indicators
            deduped_records = []
            if records:
                unique_records = {}
                for rec in records:
                    key = (rec[0], rec[1], rec[2])  # Still dedupe by country_id, indicator_id, survey_year
                    unique_records[key] = rec
                deduped_records = list(unique_records.values())
                print(f"✓ Prepared {len(deduped_records)} unique country indicator records")
                
                # Final verification before inserting country indicators
                print(f"🔍 Final check - Country ID {country_id} exists before inserting indicators...")
                final_check_cur = conn.cursor()
                final_check_cur.execute(f"SELECT country_id FROM {SCHEMA}.countries WHERE country_id = %s", (country_id,))
                final_country_check = final_check_cur.fetchone()
                if not final_country_check:
                    # Show all countries for debugging
                    final_check_cur.execute(f"SELECT country_id, name FROM {SCHEMA}.countries ORDER BY country_id")
                    all_countries = final_check_cur.fetchall()
                    final_check_cur.close()
                    raise ValueError(f"Country ID {country_id} missing before indicator insertion! Available: {all_countries}")
                final_check_cur.close()
                
                # Use a fresh cursor for insertion to avoid isolation issues
                insert_cur = conn.cursor()
                insert_country_indicators(insert_cur, deduped_records)
                insert_cur.close()
                conn.commit()

            # Logging only for first sheet
            if sheet_idx == 0:
                print(f"\n=== Processing first sheet: {country_name} ===")
                print("Categories and Indicators Detected:\n")
                for cat, inds in detected_indicators.items():
                    print(f"Category: {cat}")
                    for ind_name, ind_value in inds:
                        print(f"   - {ind_name}: {ind_value}")
                print("\n========================================")

            print(f"✓ Completed {country_name}: {len(deduped_records)} indicators processed")

        cur.close()
        print("\n✅ ETL process completed successfully!")
        
    except Exception as e:
        print(f"❌ Error in ETL process: {e}")
        conn.rollback()
        raise

conn = connect_db()
try:
    process_excel(EXCEL_PATH, conn)

    # Sanity check: count indicators per category
    sanity_query = f"""
        SELECT c.name AS country, ic.cat_name, COUNT(*) AS num_indicators
        FROM {SCHEMA}.country_indicators ci
        JOIN {SCHEMA}.indicators i ON ci.indicator_id = i.indicator_id
        JOIN {SCHEMA}.indicator_categories ic ON i.category_id = ic.category_id
        JOIN {SCHEMA}.countries c ON ci.country_id = c.country_id
        GROUP BY c.name, ic.cat_name
        ORDER BY c.name, ic.cat_name;
    """
    df_check = pd.read_sql(sanity_query, conn)
    display(df_check)

finally:
    conn.close()